import random
import sys
import re
import os
from datetime import datetime
from openpyxl import load_workbook

"""
To run this script, you need to pass the path to the text file containing the plant lines as the first argument:
python generate_csvs.py <input_file>

This script produces two CSV files with the specified format:
- [mo]_[dd]_[year]_Mapping_Population_R1.csv    for biological replicate 1
- [mo]_[dd]_[year]_Mapping_Population_R2.csv    for biological replicate 2

The idea is to feed these datasheets directly into the Blackbird.

Written by Michele Wiseman of Oregon State University
April 13, 2023
Version 1.0
"""

# Save current working directory for printing later
cwd = os.getcwd()

# Read plant lines from the text file
input_file = sys.argv[1]
with open(input_file, 'r') as f:
    plant_lines = [line.strip() for line in f.readlines()]

# Check if plant lines contain invalid characters
for line in plant_lines:
    if not re.match(r'^[A-Za-z0-9_-]*$', line):
        print(f"Error: Invalid character(s) found in line '{line}'. Only A-Z, a-z, 0-9, and '_' are allowed.")
        sys.exit(1)

# Randomize order and create three replicates of each plant line
random.shuffle(plant_lines)
tech_reps = range(3) 
plant_lines_replicates = [line for line in plant_lines for _ in tech_reps]

# Helper function to create the data in the specified format
def create_data(lines, suffix):
    data = []
    for i in range(100):
        row = []
        for j in range(4):
            index = i + j * 100
            if index > 350:
                break
            if index < len(lines):
                row.extend([f'{i + 1 + j * 100}', f'{lines[index]}_{suffix}'])
            else:
                row.extend([f'{i + 1 + j * 100}', None])
        data.append(row)
    return data

# Create data for R1 and R2
data_R1 = create_data(plant_lines_replicates, 'R1')
data_R2 = create_data(plant_lines_replicates, 'R2')

# Generate the current date string
date_str = datetime.now().strftime('%m_%d_%Y')
today = datetime.today()
today_mm = today.strftime('%m')
today_dd = today.strftime('%d')
today_yy = today.strftime('%Y')


# Save R1 and R2 data in the template .xlsx file with the specified format
for idx, data in enumerate([data_R1, data_R2]):
    # Load the template workbook and select the active worksheet
    wb = load_workbook('labels_template.xlsx')
    ws = wb.active

    ws.cell(row=1, column=4).value = 'Mapping_Population_2023'
    ws.cell(row=1, column=8).value = idx + 1

    ws.cell(row=2, column=4).value = int(today_mm)
    ws.cell(row=2, column=6).value = int(today_dd)
    ws.cell(row=2, column=8).value = int(today_yy)

    # Fill the template with data
    for row_idx, row_data in enumerate(data, start=3):
        for col_idx, cell_data in enumerate(row_data):
            ws.cell(row=row_idx, column=col_idx+1).value = cell_data

    # Save the filled template with the specified naming convention
    wb.save(f'{date_str}_Mapping_Population_R{idx+1}.xlsx')

# Print statements to confirm the script ran successfully
print(f"Number of plant lines: {len(plant_lines)}")
print(f"Number of technical replicates: {len(tech_reps)}")
print("""
Datasheets saved as:
{cwd}/{r1_filename}
{cwd}/{r2_filename}
""".format(cwd=cwd, r1_filename=f'{date_str}_Mapping_Population_R1.xlsx', r2_filename=f'{date_str}_Mapping_Population_R2.xlsx'))

# To add: argparse for technical and bioreps and a flag to specify the output directory (pending feedback from Dani)
